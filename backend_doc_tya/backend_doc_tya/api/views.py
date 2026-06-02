# api/views.py
# from datetime import timezone
import io
import re
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.db.models import Min, F, Q  # ← Импортируем агрегатные функции
from django.contrib.postgres.aggregates import ArrayAgg  # ← Для PostgreSQL
from openpyxl.descriptors import Min
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path
from django.conf import settings
from rest_framework import viewsets, status
from .models import Enrollment, Student, Module, Certificate, Staff, Section, SubSection, GroupSchedule, \
    LESSON_TIME_CHOICES, SectionResult, ScheduleItem, ScheduleStatus
from django.db import transaction
from rest_framework.permissions import AllowAny
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate
import logging
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.db.utils import IntegrityError
from .models import Module, Student, License
from .serializers import EnrollmentCreateSerializer, ModuleSelectSerializer, StudentSearchSerializer, \
    EnrollmentJournalSerializer, CertificateEnrollmentSerializer, EnrollmentUpdateSerializer, EnrollmentListSerializer, \
    CertificatePreviewSerializer, CertificateCreateSerializer, StudentCreateSerializer, \
    EnrollmentSerializer, GroupScheduleSerializer, GroupScheduleCreateSerializer, ScheduleItemBulkUpdateSerializer
from .utils.import_excel import import_students_excel
from .utils.rauc_export import generate_rauc_excel
from api.models import Enrollment
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404
from django.template.loader import render_to_string
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from io import BytesIO
from django.db.models import Prefetch,Min, Max
import os
import json
from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import utils as xl_utils
from api.utils.folders import ensure_group_folder
import logging
from api.utils.schedule_export import generate_schedule_excel
from .models import Enrollment
from .utils.time_formatter import format_detailed_time_block
from django.http import JsonResponse  # Или from rest_framework.response import Response
from django.db.models import Min, F
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from .models import Enrollment

from .utils.schedule_constants import SLOT_LOOKUP
from rest_framework import viewsets, status, decorators, exceptions
from rest_framework.response import Response
from django.db import transaction
from django.shortcuts import get_object_or_404
from django.http import HttpResponse, HttpResponseBadRequest
from .models import GroupSchedule
from .serializers import (
    GenerateScheduleInputSerializer, RegenerateScheduleInputSerializer,
    ScheduleItemUpdateInputSerializer, ScheduleItemBulkUpdateInputSerializer,
    FinalizeScheduleInputSerializer,
)
from django.template.loader import render_to_string
from api.utils.schedule_calc import get_schedule_times




logger = logging.getLogger(__name__)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_modules(request):
    modules = Module.objects.select_related('course').order_by('code', 'title')
    return Response(ModuleSelectSerializer(modules, many=True).data)

# =============================================================================
# СЛУШАТЕЛИ
# =============================================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_student(request):
    serializer = StudentCreateSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response({'message': '✅ Слушатель успешно создан'}, status=201)
    return Response(serializer.errors, status=400)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def search_students(request):

    q = request.query_params.get('q', '').strip()

    # 🔹 Минимальная длина для поиска
    if len(q) < 2:
        return Response([])

    # 🔹 Ищем по началу фамилии, имени, отчества, СНИЛС, табельному (регистронезависимо)
    results = Student.objects.filter(
        Q(surname__istartswith=q) |
        Q(name__istartswith=q) |
        Q(patronymic__istartswith=q) |
        Q(snils__icontains=q) |
        Q(employee_id__icontains=q)
    ).values(
        'id', 'surname', 'name', 'patronymic',
        'employee_id', 'snils', 'dcat_id'
    ).order_by('surname', 'name')[:10]  # 🔹 Ограничиваем 10 результатами

    # 🔹 Добавляем display_name для удобства фронта
    for r in results:
        r['display_name'] = f"{r['surname']} {r['name']} {r['patronymic'] or ''}".strip()

    return Response(list(results))


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_students_view(request):
    """POST /api/students/import-excel/ — Импорт слушателей из Excel"""

    if 'file' not in request.FILES:
        return Response({'error': 'Файл не передан'}, status=400)

    excel_file = request.FILES['file']

    # 🔹 Проверка расширения
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        return Response({'error': 'Поддерживаются только файлы .xlsx или .xls'}, status=400)

    try:
        created, updated, errors, error_details = import_students_excel(excel_file)

        return Response({
            'message': 'Импорт завершён',
            'summary': {
                'created': created,
                'updated': updated,
                'errors': errors
            },
            'error_details': error_details[:20]  # показываем первые 20 ошибок
        }, status=200)

    except Exception as e:
        return Response({'error': f'Ошибка обработки файла: {str(e)}'}, status=500)


# =============================================================================
# ЗАЧИСЛЕНИЯ
# =============================================================================


# api/views.py (или отдельный файл для ViewSet)
from rest_framework import viewsets, permissions
from rest_framework.response import Response
from django.db.models import F
from .models import Enrollment
from .serializers import EnrollmentListSerializer  # создайте простой сериализатор


class EnrollmentViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Только чтение: список и детали зачислений.
    Поддерживает фильтрацию ?module_id=...&status=...
    """
    queryset = Enrollment.objects.select_related('module', 'student').all()
    serializer_class = EnrollmentListSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        # 🔹 Фильтр по модулю
        module_id = self.request.query_params.get('module_id')
        if module_id:
            qs = qs.filter(module_id=module_id)
        # 🔹 Фильтр по статусу
        status = self.request.query_params.get('status')
        if status:
            qs = qs.filter(status=status)
        return qs

    # 🔹 Экшен для получения уникальных групп (для селектора)
    def list_groups(self, request):
        groups = self.queryset.values(
            'id', 'group', 'module_id',
            module_code=F('module__code'),
            module_title=F('module__title')
        ).distinct('group').order_by('group')
        return Response(list(groups))
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def enrollment_import_excel(request):
    mode = request.data.get('mode', 'preview')
    file = request.FILES.get('file')
    if not file:
        return Response({'error': 'Файл не загружен'}, status=400)

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception as e:
        return Response({'error': f'Ошибка чтения Excel: {e}'}, status=400)

    staff = getattr(request.user, 'staff_profile', None)
    if not staff:
        return Response({'error': 'Нет профиля сотрудника'}, status=403)

    # 🔹 Нормализация кода модуля (убираем артефакты Excel)
    def normalize_code(val):
        if not val: return ''
        s = str(val).strip()
        # 1. Убираем невидимые символы Excel
        s = re.sub(r'[\s\u00A0\u200B\uFEFF]+', ' ', s)
        # 2. Все виды тире/дефисов → стандартный '-'
        s = re.sub(r'[\u2013\u2014\u2212\-]+', '-', s)
        s = re.sub(r'\s*-\s*', '-', s)
        # 3. 🔑 Заменяем латинские визуальные двойники на кириллицу
        s = s.translate(str.maketrans(
            "MmCcOoAaEeKkPpXxYyTtHhBb",
            "МмСсОоАаЕеКкРрХхУуТтНнВв"
        ))
        return s.upper()

    # 🔹 Кэшируем ТОЛЬКО по полю code
    modules_by_code = {}
    for m in Module.objects.all():
        norm = normalize_code(m.code)
        if norm:
            modules_by_code[norm] = m

    results = []
    valid_data = []

    for idx, row in enumerate(rows, start=2):
        if not any(row): continue
        try:
            app, mod_code, grp, enr_date, start_date, comp_date, loc, num_in_grp, snils, sdo_date = row[:10]

            # 1. Студент
            snils_clean = str(snils).replace(' ', '').replace('-', '')
            students_by_snils = {s.snils.replace(' ', '').replace('-', ''): s for s in Student.objects.all()} # ⚠️ лучше вынести за цикл, но оставлю как у тебя для минимальных правок
            student = students_by_snils.get(snils_clean)
            if not student:
                raise ValueError(f"Слушатель с СНИЛС '{snils_clean}' не найден")

            # 2. 🔍 Поиск модуля СТРОГО по code
            m_str = normalize_code(mod_code)
            module = modules_by_code.get(m_str)
            if not module:
                raise ValueError(f"Модуль '{mod_code}' не найден в БД. (Нормализовано: '{m_str}')")

            # 3. Даты
            def parse_date(val, fallback=None):
                if isinstance(val, datetime): return val.date()
                if isinstance(val, (int, float)) and val > 30000:
                    return datetime(1899, 12, 30).date() + timedelta(days=int(val))
                if val:
                    for fmt in ('%d/%m/%y', '%Y-%m-%d', '%d.%m.%Y'):
                        try: return datetime.strptime(str(val).strip(), fmt).date()
                        except ValueError: continue
                return fallback

            loc_clean = str(loc).strip().upper()
            if loc_clean not in ['KJA', 'DME']:
                raise ValueError(f"Локация '{loc}' недопустима. Используйте KJA или DME")

            # 4. Проверка дублей (твоя логика без изменений)
            existing_enr = Enrollment.objects.filter(student=student, module=module).order_by('-enrolled_at').first()
            if existing_enr:
                if existing_enr.status == 'completed' and existing_enr.completed_at:
                    days_since = (timezone.now().date() - existing_enr.completed_at).days
                    VALIDITY_DAYS = 365
                    if days_since < VALIDITY_DAYS:
                        raise ValueError(f"Модуль успешно пройден {days_since} дн. назад. Повторное зачисление доступно через {VALIDITY_DAYS - days_since} дн.")
                if existing_enr.status in ['enrolled', 'in_progress']:
                    raise ValueError(f"Модуль уже в статусе «{existing_enr.status}». Завершите или отмените текущее зачисление.")

            # 5. Сборка данных
            data = {
                'student': student, 'module': module, 'status': 'enrolled',
                'group': str(grp).strip(), 'application': str(app).strip(),
                'number_in_group': int(num_in_grp),
                'enrolled_at': parse_date(enr_date, timezone.now().date()),
                'start_face_to_face': parse_date(start_date, timezone.now().date()),
                'completed_at': parse_date(comp_date),
                'location': loc_clean, 'enrolled_by': staff,
                'start_sdo': parse_date(sdo_date),

            }
            valid_data.append(data)
            results.append({
                'row': idx,
                'status': 'ok',
                'student': f"{student.surname} {student.name} {student.patronymic or ''}".strip()
            })

        except Exception as e:
            results.append({'row': idx, 'status': 'error', 'error': str(e)})

    if mode == 'preview':
        return Response({'results': results, 'valid_count': len(valid_data)})
    else:
        created = 0
        for d in valid_data:
            Enrollment.objects.create(**d)
            created += 1
        return Response({'created': created, 'total_valid': len(valid_data)})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reenroll_enrollment(request, pk):
    """POST /api/enrollments/<pk>/reenroll/ — Перезачисление на модуль"""
    old = get_object_or_404(Enrollment, id=pk)

    # 🔹 1. Проверка статуса
    if old.status in ['enrolled', 'in_progress']:
        return Response({'error': 'Нельзя перезачислить активную запись. Завершите или отмените текущее зачисление.'},
                        status=400)

    # 🔹 2. Проверка срока действия (только для успешно завершённых)
    if old.status == 'completed' and old.completed_at:
        days_since = (timezone.now().date() - old.completed_at).days
        VALIDITY_DAYS = 365  # 🔹 Период действия сертификата

        if days_since < VALIDITY_DAYS:
            return Response({
                'error': f'Сертификат действует ещё {VALIDITY_DAYS - days_since} дн. Повторное зачисление заблокировано.'
            }, status=400)

    # 🔹 3. Создание нового зачисления
    staff = getattr(request.user, 'staff_profile', None)
    new_enrollment = Enrollment.objects.create(
        student=old.student,
        module=old.module,
        status='enrolled',
        group=old.group,
        application=old.application,
        number_in_group=old.number_in_group,
        enrolled_at=timezone.now().date(),
        location=old.location,
        enrolled_by=staff,

    )

    # 🔹 Опционально: пометить старую запись как архивную
    # old.status = 'archived'  # (добавьте 'archived' в STATUS_CHOICES модели, если нужно)
    # old.save(update_fields=['status'])

    return Response({
        'message': '✅ Слушатель успешно перезачислен',
        'new_enrollment_id': new_enrollment.id
    }, status=201)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_enrollment(request):
    """POST /api/enrollments/ — Создание зачисления (и папки группы при необходимости)"""

    logger.info(f"[Enrollment] Request data: {request.data}")

    serializer = EnrollmentCreateSerializer(data=request.data, context={'request': request})

    if serializer.is_valid():
        try:
            # 🔹 1. Сохраняем зачисление в БД
            instance = serializer.save()

            # 🔹 2. Получаем код группы из сохранённого объекта
            group_code = instance.group

            # 🔹 3. Создаём папку для группы (если первая запись с таким кодом)
            if group_code:
                folder_path = ensure_group_folder(group_code)
                logger.info(f"📁 Папка группы '{group_code}': {folder_path}")

            logger.info(f"[Enrollment] Created: {instance.id}")
            return Response({
                'success': True,
                'message': 'Зачисление создано',
                'id': instance.id,
                'group_folder': folder_path if group_code else None
            }, status=201)

        except IntegrityError as e:
            logger.warning(f"[Enrollment] IntegrityError: {e}")
            return Response({'error': 'Слушатель уже зачислен на этот модуль'}, status=400)
        except Exception as e:
            logger.error(f"[Enrollment] Unexpected error: {e}")
            return Response({'error': 'Внутренняя ошибка сервера'}, status=500)

    logger.warning(f"[Enrollment] Validation errors: {serializer.errors}")
    return Response(serializer.errors, status=400)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def enrollment_list(request):
    """
    GET /api/enrollments/list/?status__in=enrolled,in_progress,completed&module_code=...
    Возвращает отфильтрованные назначения (поддерживает множественный выбор статуса)
    """
    qs = Enrollment.objects.select_related(
        'student', 'module', 'module__course'
    ).prefetch_related('certificates', 'section_results')

    # 🔹 FIX: Фильтр по статусу — поддерживаем status__in
    status_param = request.query_params.get('status__in')
    if status_param:
        # Ожидаем: "enrolled,in_progress,completed"
        statuses = [s.strip() for s in status_param.split(',') if s.strip()]
        valid_statuses = [s for s in statuses if s in ['enrolled', 'in_progress', 'completed', 'failed']]
        if valid_statuses:
            qs = qs.filter(status__in=valid_statuses)
    else:
        # Если параметр не передан — по умолчанию показываем активные (не failed)
        qs = qs.exclude(status='failed')



    # 🔹 Фильтр по дате завершения (опционально)
    completed_after = request.query_params.get('completed_after')
    if completed_after:
        qs = qs.filter(completed_at__gte=completed_after)

    # 🔹 Фильтр по модулю
    module_code = request.query_params.get('module_code')
    if module_code:
        code_clean = module_code.strip()
        qs = qs.filter(
            Q(module__code__iexact=code_clean) |
            Q(module__code__icontains=code_clean)
        )

    # 🔹 Если не админ — только свои назначения (или те, что создал этот staff)
    if not request.user.is_staff:
        staff = getattr(request.user, 'staff_profile', None)
        if staff:
            # Показываем назначения, которые сделал этот сотрудник, ИЛИ где он оценивал разделы
            qs = qs.filter(
                Q(enrolled_by=staff) |
                Q(section_results__evaluator=staff)
            ).distinct()

    serializer = EnrollmentListSerializer(qs.order_by('-completed_at', '-id'), many=True)
    return Response(serializer.data, status=200)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_enrollment_groups(request):
    """
    Возвращает список групп из активных назначений с привязанными модулями.
    ?module_id=XXX — опциональная фильтрация по модулю
    """
    module_id = request.query_params.get('module_id')

    qs = Enrollment.objects.select_related('module', 'student').filter(
        status__in=['enrolled', 'in_progress']
    ).order_by('-enrolled_at')

    if module_id:
        qs = qs.filter(module_id=module_id)

    # Группируем: одна запись на уникальную пару (группа, модуль)
    seen = set()
    groups = []
    for enr in qs:
        key = (enr.group, enr.module_id)
        if key not in seen:
            seen.add(key)
            groups.append({
                'group': enr.group,
                'module_id': enr.module.id,
                'module_code': enr.module.code,
                'module_title': enr.module.title,
                'enrollment_id': enr.id,  # 🔹 Для генерации расписания
                'start_face_to_face': enr.start_face_to_face,
                'start_sdo': enr.start_sdo,
            })

    return Response(groups)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def current_user(request):
    """Данные текущего сотрудника"""
    user = request.user
    staff = getattr(user, 'staff_profile', None)
    return Response({
        'id': user.id,
        'username': user.username,
        'name': staff.name if staff else user.username,
        'position': staff.position if staff else '',
    })


@api_view(['POST'])
@permission_classes([AllowAny])
def staff_login(request):
    # 🪵 Логируем входящие данные
    logger.info(f"[Login] Request data: {request.data}")
    logger.info(f"[Login] Content-Type: {request.content_type}")

    # 🔁 Принимаем оба варианта: username или rauts_id
    username = request.data.get('username') or request.data.get('rauts_id')
    password = request.data.get('password')

    if not username or not password:
        logger.warning(f"[Login] Missing fields: username={username}, password={'***' if password else None}")
        return Response({'error': 'Логин и пароль обязательны'}, status=400)

    user = authenticate(username=str(username).strip(), password=password)

    if user:
        refresh = RefreshToken.for_user(user)
        staff_profile = getattr(user, 'staff_profile', None)
        return Response({
            'success': True,
            'access': str(refresh.access_token),
            'user': {
                'id': user.id,
                'username': user.username,
                'name': staff_profile.name if staff_profile else user.username,
            }
        })

    logger.warning(f"[Login] Auth failed for username: {username}")
    return Response({'error': 'Неверный логин или пароль'}, status=401)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def journal_enrollments_list(request):
    """
    Журнал зачислений: показываем записи, которые ЕЩЁ не завершены окончательно.
    Запись уходит из журнала, только когда:
    - status == 'completed' И total_mark > 0
    """
    enrollments = Enrollment.objects.filter(
        # Показываем, если:
        # 1. Статус НЕ "завершен"
        # ИЛИ
        # 2. Нет оценки (null или 0)
        ~Q(status='completed') | Q(total_mark__isnull=True) | Q(total_mark=0)
    ).select_related('student', 'module', 'enrolled_by').order_by('group', '-enrolled_at')

    return Response(EnrollmentJournalSerializer(enrollments, many=True).data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def journal_structure(request, enrollment_id):
    """GET /api/journals/<enrollment_id>/structure/"""

    # 🔹 ПРАВИЛЬНЫЙ путь с учётом related_name
    enrollment = get_object_or_404(
        Enrollment.objects.select_related('student', 'module').prefetch_related(
            'module__stages__sections',  # ✅ Module → stages → sections
            'section_results__section__stage'  # ✅ Для существующих результатов
        ),
        id=enrollment_id
    )

    # Существующие результаты
    existing_results = {
        res.section_id: {
            'id': res.id,
            'hours_completed': float(res.hours_completed),
            'grade': float(res.grade) if res.grade else None,
            'is_passed': res.is_passed,
            'completed_at': res.completed_at.isoformat() if res.completed_at else None,
            'ngroup': res.ngroup
        }
        for res in enrollment.section_results.all()  # ✅ Использует кэш prefetch
    }

    serializer = EnrollmentJournalSerializer(enrollment)
    data = serializer.data
    data['existing_results'] = existing_results

    return Response(data)


# 🔹 ИСПРАВЛЕННАЯ вьюха
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def journal_results_save(request):
    """POST /api/journals/results/ — Сохраняет массив результатов разделов"""

    staff = getattr(request.user, 'staff_profile', None)
    if not staff:
        return Response({'error': 'Только для сотрудников'}, status=403)

    enrollment_id = request.data.get('enrollment_id')
    results = request.data.get('results', [])

    if not enrollment_id or not results:
        return Response({'error': 'Укажите enrollment_id и результаты'}, status=400)

    enrollment = get_object_or_404(Enrollment, id=enrollment_id)
    saved_count = 0
    errors = []

    for item in results:
        section_id = item.get('section')
        if not section_id:
            continue

        # 🔹 FIX: создаём defaults БЕЗ полей 'section' и 'enrollment'
        defaults = {
            'hours_completed': float(item.get('hours_completed', 0)),
            'grade': float(item['grade']) if item.get('grade') is not None else None,
            'is_passed': item.get('is_passed'),
            'completed_at': item.get('completed_at'),
            'evaluator': staff,  # текущий сотрудник
            'ngroup': item.get('ngroup') or f"{enrollment.group}-{enrollment.application}",
        }

        # 🔹 Валидация числовых полей
        try:
            if defaults['grade'] is not None:
                defaults['grade'] = float(defaults['grade'])
            defaults['hours_completed'] = float(defaults['hours_completed'])
        except (ValueError, TypeError) as e:
            errors.append(f"Section {section_id}: invalid number - {e}")
            continue

        # 🔹 Валидация по типу оценки
        try:
            section = Section.objects.get(id=section_id)
            if section.grade_type == 'binary' and defaults['is_passed'] is None:
                raise ValueError('is_passed required for binary')
            if section.grade_type == 'numeric' and defaults['grade'] is None:
                raise ValueError('grade required for numeric')
        except Section.DoesNotExist:
            errors.append(f"Section {section_id}: not found")
            continue
        except ValueError as e:
            errors.append(f"Section {section_id}: {e}")
            continue

        try:
            # 🔹 FIX: используем section_id в фильтрах И в defaults
            SectionResult.objects.update_or_create(
                enrollment=enrollment,
                section_id=section_id,  # ✅ Фильтр по ID
                defaults={
                    **defaults,
                    'section_id': section_id,  # ✅ В defaults тоже _id!
                }
            )
            saved_count += 1
        except Exception as e:
            errors.append(f"Section {section_id}: {str(e)}")

    if errors:
        return Response({
            'message': f'Сохранено: {saved_count}, Ошибок: {len(errors)}',
            'errors': errors[:5]
        }, status=207)

    return Response({
        'message': f'✅ Сохранено результатов: {saved_count}',
        'enrollment_id': enrollment_id
    }, status=201 if saved_count > 0 else 200)



@api_view(['GET', 'PATCH'])
@permission_classes([IsAuthenticated])
def enrollment_journal_detail(request, pk):
    """Получение записи или её обновление"""
    try:
        enrollment = Enrollment.objects.select_related('student', 'module').get(pk=pk)
    except Enrollment.DoesNotExist:
        return Response({'error': 'Запись не найдена'}, status=404)

    if request.method == 'GET':
        serializer = EnrollmentJournalSerializer(enrollment)
        return Response(serializer.data)

    # PATCH: обновление
    serializer = EnrollmentUpdateSerializer(
        enrollment,
        data=request.data,
        partial=True,
        context={'request': request}
    )

    if serializer.is_valid():
        updated_instance = serializer.save()
        # 🔹 Возвращаем данные в формате "чтения", чтобы фронт обновил UI
        return Response(EnrollmentJournalSerializer(updated_instance).data)

    # 🔹 Возвращаем ошибки валидации
    return Response(serializer.errors, status=400)

@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def enrollment_update_status(request, pk):
    """
    PATCH /api/enrollments/<pk>/
    Обновление статуса зачисления (только поле status)
    """
    staff = getattr(request.user, 'staff_profile', None)
    if not staff or not staff.fptitle:
        return Response({'error': 'Нет прав'}, status=403)

    enrollment = get_object_or_404(Enrollment, id=pk)

    # 🔹 Разрешаем менять только status
    new_status = request.data.get('status')
    valid_statuses = ['enrolled', 'in_progress', 'completed', 'failed']

    if new_status not in valid_statuses:
        return Response({'error': f'Неверный статус. Допустимы: {valid_statuses}'}, status=400)

    enrollment.status = new_status

    # 🔹 Авто-заполнение completed_at при завершении
    if new_status == 'completed' and not enrollment.completed_at:
        enrollment.completed_at = timezone.now().date()

    enrollment.save(update_fields=['status', 'completed_at'])

    return Response({
        'id': enrollment.id,
        'status': enrollment.status,
        'completed_at': enrollment.completed_at
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_journal_html(request):
    """Генерация HTML журнала"""
    ids = request.data.get('enrollments_ids', [])
    enrollments = Enrollment.objects.filter(id__in=ids).select_related('student', 'module')

    html_content = render_to_string('reports/journal_template.html', {
        'enrollments': enrollments,
        'generated_at': timezone.now(),
        'staff_name': request.user.staff_profile.name if hasattr(request.user, 'staff_profile') else ''
    })

    response = HttpResponse(html_content, content_type='text/html; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="journal.html"'
    return response



# =============================================================================
# СЕРТИФИКАТЫ — УНИВЕРСАЛЬНЫЙ ЭНДПОИНТ
# =============================================================================

# 🔹 Маппинг: код модуля → шаблон контента (наследуется от course_base.html)
CERTIFICATE_CONTENT_TEMPLATES = {
    "ППП.АУЦ.11 - М.1": "certificates/modules/c11/content_c11_m1.html",
    "ППП.АУЦ.11 - М.2": "certificates/modules/c11/content_c11_m2.html",
    "ППП.АУЦ.02 - М.1": "certificates/modules/c02/content_c02_m1.html",
    "ППП.АУЦ.02 - М.2-NG": "certificates/modules/c02/content_c02_m2_ng.html",
    "ППП.АУЦ.02 - М.2-CL": "certificates/modules/c02/content_c02_m2_cl.html",
    "ППП.АУЦ.02 - М.3": "certificates/modules/c02/content_c02_m3.html",
    "ППП.АУЦ.02 - М.4": "certificates/modules/c02/content_c02_m4.html",
    "ППП.АУЦ.02 - М.5": "certificates/modules/c02/content_c02_m5.html",
    "ППП.АУЦ.02 - М.6": "certificates/modules/c02/content_c02_m6.html",
    "ППП.АУЦ.02 - М.7": "certificates/modules/c02/content_c02_m7.html",

    # Добавьте другие модули по мере необходимости
}
DEFAULT_CONTENT_TEMPLATE = "certificates/content_default.html"
BASE_TEMPLATE = "certificates/course_base.html"  # Общая оболочка


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def certificate_handler(request, pk=None):
    """
    GET /api/certificates/<pk>/ — предпросмотр сертификата
    POST /api/certificates/ — генерация + сохранение в БД + файл
    """

    def _get_content_template_name(enrollment):
        """Возвращает имя шаблона контента в зависимости от модуля"""
        module_code = getattr(enrollment.module, "code", "")
        return CERTIFICATE_CONTENT_TEMPLATES.get(module_code, DEFAULT_CONTENT_TEMPLATE)

    def _build_context(enrollment, staff_profile, approver_name=None):
        module = enrollment.module
        course = getattr(module, "course", None)

        # 🔹 Безопасное получение лицензии из БД
        license_obj = License.objects.first()
        if license_obj:
            license_data = {
                'license_number': getattr(license_obj, 'license_number', '_______'),
                'license_date': getattr(license_obj, 'license_date', '_______'),
                'favt_cert_number': getattr(license_obj, 'favt_cert_number', '_______'),
                'favt_cert_date': getattr(license_obj, 'favt_cert_date', '_______'),
            }
        else:
            license_data = {
                'license_number': '_______',
                'license_date': '_______',
                'favt_cert_number': '_______',
                'favt_cert_date': '_______',
            }

        return {
            "enrollment": enrollment,
            "student": enrollment.student,
            "module": module,
            "course": course,  # Добавляем весь объект course для гибкости
            "course_name": getattr(course, "name", "_______"),
            "total_hours": getattr(module, "total_hours", None) or "___",
            "elearning_hours": getattr(module, "elearning_hours", None) or "___",
            "license": license_data,
            "cert_number": f"{enrollment.group}-{enrollment.application}-{enrollment.number_in_group}",
            "issued_at": timezone.now().strftime("%d.%m.%Y"),
            "prepared_by": staff_profile.name if staff_profile else "_______",
            "approved_by": approver_name or "_______",
            "grade": enrollment.total_mark,
            "completion_date": enrollment.completed_at,
            "location_name": enrollment.get_location_display(),
            # 🔹 Для отладки: дамп контекста (убрать в продакшене)
            "context_dump": None,  # Можно заполнить при debug=True
            "debug": getattr(settings, 'DEBUG', False),
            "for_pdf": False,  # Флаг для скрытия кнопок при генерации PDF
        }

    # ========================================================================
    # 🔹 GET: Предпросмотр (без сохранения в БД)
    # ========================================================================
    if request.method == "GET" and pk:
        try:
            enrollment = Enrollment.objects.select_related("student", "module").get(
                pk=pk, status="completed", total_mark__gt=0
            )
        except Enrollment.DoesNotExist:
            return Response({"error": "Запись не найдена или не завершена"}, status=404)

        staff = getattr(request.user, "staff_profile", None)
        # 🔹 Передаём абсолютный корень сайта
        # ✅ ВЫЗЫВАЕМ ФУНКЦИЮ, а не собираем вручную!
        context = _build_context(
            enrollment=enrollment,
            staff_profile=staff,
            approver_name="Директор АУЦ",  # заглушка для превью
        )

        # 🔹 Добавляем base_url только для GET (нужен для загрузки стилей/картинок)
        context["base_url"] = request.build_absolute_uri('/')



        content_template = _get_content_template_name(enrollment)  # ← используем ту же функцию, что в POST
        html = render_to_string(content_template, context, request=request)
        return HttpResponse(html, content_type="text/html; charset=utf-8")

    # ========================================================================
    # 🔹 POST: Генерация + Сохранение в БД
    # ========================================================================
    if request.method == "POST":
        staff = getattr(request.user, "staff_profile", None)
        if not staff or not getattr(staff, "fptitle", False):
            return Response({"error": "Нет прав на оформление (fptitle=False)"}, status=403)

        enrollment_id = request.data.get("enrollment_id")
        cert_type = request.data.get("cert_type", "successful")
        approved_by_id = request.data.get("approved_by_id")

        if not enrollment_id:
            return Response({"error": "Укажите enrollment_id"}, status=400)

        # 🔹 Гибкая проверка: разрешены только модули из маппинга
        allowed_codes = list(CERTIFICATE_CONTENT_TEMPLATES.keys())
        try:
            enrollment = Enrollment.objects.select_related("student", "module").get(
                id=enrollment_id,
                module__code__in=allowed_codes,  # ← теперь не хардкод!
                status="completed",
                total_mark__gt=0,
                completed_at__isnull=False
            )
        except Enrollment.DoesNotExist:
            return Response({
                "error": f"Зачисление не найдено или не относится к поддерживаемым модулям: {', '.join(allowed_codes)}"
            }, status=404)

        # 🔄 Проверка дубликата
        if Certificate.objects.filter(enrollment=enrollment, cert_type=cert_type).exists():
            cert = Certificate.objects.get(enrollment=enrollment, cert_type=cert_type)
            return Response({
                "message": "Сертификат уже создан",
                "cert_number": cert.cert_number,
                "file_url": f"{settings.MEDIA_URL}{cert.file_path}"
            }, status=200)

        # 👮 Определение утверждающего (tptitle=True)
        approver = None
        if approved_by_id:
            approver = Staff.objects.filter(id=approved_by_id, tptitle=True, is_active=True).first()
        elif getattr(staff, "tptitle", False):
            approver = staff
        else:
            approver = Staff.objects.filter(tptitle=True, is_active=True).first()

        if not approver:
            return Response({"error": "В системе нет сотрудника с правом подписи (tptitle)"}, status=400)

        # 📝 Рендер шаблона
        context = _build_context(enrollment, staff, approver.name)
        context["for_pdf"] = True  # Скрываем кнопки в сохраняемой версии

        content_template = _get_content_template_name(enrollment)
        html_content = render_to_string(content_template, context, request=request)

        # 💾 Сохранение файла
        cert_number = context["cert_number"]
        safe_cert_number = cert_number.replace("/", "_").replace(" ", "_").replace("\\", "_")
        upload_subpath = f"certificates/{timezone.now().strftime('%Y/%m')}"
        media_root = Path(settings.MEDIA_ROOT).resolve()
        upload_dir = media_root / upload_subpath
        upload_dir.mkdir(parents=True, exist_ok=True)

        filename = f"cert_{safe_cert_number}_{timezone.now().strftime('%Y%m%d_%H%M')}.html"
        abs_file_path = upload_dir / filename

        with open(abs_file_path, "w", encoding="utf-8") as f:
            f.write(html_content)

        rel_file_path = abs_file_path.relative_to(media_root)

        # 🗄️ Создание записи в БД
        certificate = Certificate.objects.create(
            enrollment=enrollment,
            cert_type=cert_type,
            cert_number=cert_number,
            prepared_by=staff,
            approved_by=approver,
            file_path=str(rel_file_path),
            comment=f"Модуль: {enrollment.module.code}"
        )

        logger.info(f"✅ Сертификат выписан: {cert_number} | Подготовил: {staff.name} | Утвердил: {approver.name}")

        return Response({
            "message": "Сертификат успешно создан",
            "cert_number": certificate.cert_number,
            "file_url": f"{settings.MEDIA_URL}{certificate.file_path}",
            "issued_at": certificate.issued_at,
            "results_snapshot": certificate.results_snapshot
        }, status=201)

    return Response({"error": "Неверный метод"}, status=400)

# =============================================================================
# СЕРТИФИКАТЫ НОВЫЙ ПОДХОД
# =============================================================================


# 🔹 Список зачислений, доступных для выписки сертификата
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def certificates_eligible_list(request):
    """
    GET /api/certificates/eligible/
    Возвращает зачисления, для которых можно выписать сертификат
    """
    # Только завершённые с оценкой > 0, для которых ещё нет такого типа сертификата
    qs = Enrollment.objects.select_related('student', 'module', 'module__course').filter(
        status='completed',
        total_mark__gt=0
    ).exclude(
        # Исключаем те, для которых уже есть сертификат нужного типа
        certificates__cert_type=request.query_params.get('cert_type', 'successful')
    )

    # Фильтр по модулю (опционально)
    module_code = request.query_params.get('module_code')
    if module_code:
        qs = qs.filter(module__code__icontains=module_code)

    # 🔹 Если не админ — только свои назначения
    if not request.user.is_staff:
        staff = getattr(request.user, 'staff_profile', None)
        if staff:
            qs = qs.filter(
                Q(enrolled_by=staff) | Q(section_results__evaluator=staff)
            ).distinct()

    # Сериализация (упрощённо)
    data = []
    for enr in qs.order_by('-completed_at'):
        data.append({
            'id': enr.id,
            'student_name': f"{enr.student.surname} {enr.student.name}",
            'module_code': enr.module.code,
            'module_title': enr.module.title,
            'group': enr.group,
            'application': enr.application,
            'total_mark': enr.total_mark,
            'completed_at': enr.completed_at.isoformat() if enr.completed_at else None,
            'has_certificate': enr.certificates.exists()
        })

    return Response(data)


# 🔹 Предпросмотр сертификата
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def certificate_preview(request, pk):
    """
    GET /api/certificates/<pk>/preview/
    Возвращает данные для предпросмотра сертификата
    """
    enrollment = get_object_or_404(
        Enrollment.objects.select_related('student', 'module', 'module__course'),
        id=pk,
        status='completed',
        total_mark__gt=0
    )

    # 🔹 Получаем списки подписантов
    prepared_options = Staff.objects.filter(
        fptitle=True, is_active=True
    ).values('id', 'name').order_by('name')

    approved_options = Staff.objects.filter(
        tptitle=True, is_active=True
    ).values('id', 'name').order_by('name')

    # 🔹 Создаём "фейковый" объект Certificate для сериализации
    fake_cert = Certificate(enrollment=enrollment)

    serializer = CertificatePreviewSerializer(
        fake_cert,
        context={
            'prepared_by_options': list(prepared_options),
            'approved_by_options': list(approved_options)
        }
    )

    return Response(serializer.data)


# 🔹 Список сотрудников для выбора
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def certificate_staff_list(request):
    """
    GET /api/certificates/staff/
    Возвращает списки сотрудников с правами подготовки/утверждения
    """
    return Response({
        'prepared_by': list(Staff.objects.filter(
            fptitle=True, is_active=True
        ).values('id', 'name', 'position').order_by('name')),
        'approved_by': list(Staff.objects.filter(
            tptitle=True, is_active=True
        ).values('id', 'name', 'position').order_by('name'))
    })


# 🔹 Создание сертификата

def normalize_module_code(val):
    """Та же функция нормализации, что и для импорта Excel"""
    if not val: return ''
    s = str(val).strip()
    s = re.sub(r'[\s\u00A0\u200B\uFEFF]+', ' ', s)
    s = re.sub(r'[\u2013\u2014\u2212\-]+', '-', s)
    s = re.sub(r'\s*-\s*', '-', s)
    s = s.translate(str.maketrans("MmCcOoAaEeKkPpXxYyTtHhBb", "МмСсОоАаЕеКкРрХхУуТтНнВв"))
    return s.upper()

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def certificate_create(request):
    staff = getattr(request.user, 'staff_profile', None)
    if not staff or not staff.fptitle:
        return Response({'error': 'Нет прав на оформление сертификатов'}, status=403)

    serializer = CertificateCreateSerializer(data=request.data, context={'request': request})
    if not serializer.is_valid():
        return Response(serializer.errors, status=400)

    # 🔹 1. Сохраняем сертификат
    certificate = serializer.save(prepared_by=staff)
    file_url = None

    # 🔹 Получаем все результаты
    all_results = list(certificate.enrollment.section_results
                       .select_related('section__stage')
                       .values(
        'section__title',
        'section__grade_type',
        'hours_completed',
        'grade',
        'is_passed'
    )
                       .order_by('section__stage__order', 'section__order'))

    # 🔹 Отделяем итоговую оценку по этапу
    final_exam_result = None
    section_results_filtered = []

    for r in all_results:
        if "Итоговая оценка" in r['section__title'] and "Этапа" in r['section__title']:
            final_exam_result = r
        else:
            section_results_filtered.append(r)

    try:
        # 🔹 2. Контекст для шаблона (используем ЖИВУЮ QuerySet, а не snapshot!)
        context = {
            'certificate': certificate,
            'enrollment': certificate.enrollment,
            'student': certificate.enrollment.student,
            'module': certificate.enrollment.module,
            'course': certificate.enrollment.module.course,
            'license': License.objects.first(),  # ⚠️ можно закэшировать
            'prepared_by': certificate.prepared_by,
            'approved_by': certificate.approved_by,
            'cert_number': certificate.cert_number,
            'issued_at': certificate.issued_at.strftime('%d.%m.%Y'),
            # В preview-endpoint, который отдаёт HTML (не JSON!)

            'section_results': section_results_filtered,  # ← только темы, без итога
            'final_grade': certificate.enrollment.total_mark,
            'final_exam_hours': final_exam_result['hours_completed'],
        }

        # 🔹 3. Выбор шаблона с нормализацией кода
        module_code = normalize_module_code(certificate.enrollment.module.code)

        template_map = {
            'ППП.АУЦ.11-М.1': 'certificates/modules/c11/content_c11_m1.html',
            'ППП.АУЦ.11-М.2': 'certificates/modules/c11/content_c11_m2.html',
            'ППП.АУЦ.02-М.1': 'certificates/modules/c02/content_c02_m1.html',
            'ППП.АУЦ.02-М.2-NG': 'certificates/modules/c02/content_c02_m2_ng.html',
            'ППП.АУЦ.02-М.2-CL': 'certificates/modules/c02/content_c02_m2_cl.html',
            'ППП.АУЦ.02-М.3-NG': 'certificates/modules/c02/content_c02_m3_ng.html',
            'ППП.АУЦ.02-М.3-CL': 'certificates/modules/c02/content_c02_m3_cl.html',
            'ППП.АУЦ.02-М.3-CL/NG': 'certificates/modules/c02/content_c02_m3_all.html',
            'ППП.АУЦ.02-М.4': 'certificates/modules/c02/content_c02_m4.html',
            'ППП.АУЦ.02-М.5': 'certificates/modules/c02/content_c02_m5.html',
            'ППП.АУЦ.02-М.6': 'certificates/modules/c02/content_c02_m6.html',
            'ППП.АУЦ.02-М.7': 'certificates/modules/c02/content_c02_m7.html',
        }

        template_name = template_map.get(module_code)

        if template_name:
            html_content = render_to_string(template_name, context, request=request)

            # 🔹 Сохраняем файл через Django Storage (надёжнее ручного open())
            safe_name = certificate.cert_number.replace('/', '_').replace(' ', '_')
            filename = f"cert_{safe_name}_{timezone.now().strftime('%Y%m%d_%H%M')}.html"

            # ContentFile автоматически положит файл в MEDIA_ROOT/certificates/%Y/%m/
            certificate.file_path.save(filename, ContentFile(html_content.encode('utf-8')), save=False)
            certificate.save(update_fields=['file_path'])

            file_url = request.build_absolute_uri(certificate.file_path.url)
        else:
            print(f"⚠️ Template not found for normalized module code: {module_code}")

    except Exception as e:
        print(f"❌ Certificate file generation error: {e}")
        # Запись в БД остаётся, как и задумано

    return Response({
        'message': 'Сертификат успешно создан',
        'cert_number': certificate.cert_number,
        'file_url': file_url,
        'issued_at': certificate.issued_at.isoformat() if certificate.issued_at else None,
        'id': certificate.id,
        'warning': 'Файл не сгенерирован (шаблон не найден или ошибка генерации)' if not file_url else None
    }, status=201)
# =============================================================================
# ЭКСПОРТ EXEL РАУЦ
# =============================================================================


# 🔹 Единый конфигурационный блок prefetch (используется в обоих эндпоинтах)
ENROLLMENT_PREFETCH = [
    Prefetch(
        'certificates',
        queryset=Certificate.objects.select_related('prepared_by', 'approved_by')
    ),
    Prefetch(
        'section_results',
        queryset=SectionResult.objects.select_related('section').order_by('-completed_at')
    ),
]


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def rauc_export_list(request):
    """GET /api/rauc/list/ — список зачислений для экспорта"""
    enrollments = Enrollment.objects.filter(
        status='completed',
        total_mark__isnull=False,
        total_mark__gt=0
    ).select_related(
        'student', 'module', 'module__course', 'enrolled_by'
    ).prefetch_related(*ENROLLMENT_PREFETCH)  # 🔹 Распаковка списка

    return Response(CertificateEnrollmentSerializer(enrollments, many=True).data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_rauc_excel_view(request):
    """POST /api/rauc/generate/ — генерация Excel"""
    enrollment_ids = request.data.get('enrollment_ids', [])

    if not enrollment_ids:
        return Response({'error': 'Не выбраны записи для экспорта'}, status=400)

    # 🔹 Используем тот же prefetch, что и в list
    enrollments = Enrollment.objects.filter(id__in=enrollment_ids).select_related(
        'student', 'module', 'module__course'
    ).prefetch_related(*ENROLLMENT_PREFETCH)  # 🔹 Тот же список

    if not enrollments.exists():
        return Response({'error': 'Записи не найдены'}, status=404)

    # 🔹 Функция экспорта НЕ делает своих prefetch — использует уже загруженные данные
    wb = generate_rauc_excel(enrollments)


    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"RAUC_export_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

# =============================================================================
# РАСПИСАНИЕ
# =============================================================================


class GroupScheduleViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet для работы с расписанием группы"""
    serializer_class = GroupScheduleSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return GroupSchedule.objects.select_related(
            'enrollment__module', 'curator', 'director'
        ).prefetch_related(
            'items__stage', 'items__section', 'items__instructor'
        ).all()

    @action(detail=True, methods=['patch'], url_path='update-meta')
    def update_meta(self, request, pk=None):
        """Обновление куратора и директора"""
        gs = self.get_object()

        curator_id = request.data.get('curator')
        director_id = request.data.get('director')

        if curator_id is not None:
            try:
                gs.curator_id = curator_id if curator_id else None
            except Exception:
                gs.curator = None

        if director_id is not None:
            try:
                gs.director_id = director_id if director_id else None
            except Exception:
                gs.director = None

        gs.save()

        # Возвращаем обновлённое расписание
        gs = GroupSchedule.objects.prefetch_related(
            'items__stage', 'items__section', 'items__instructor'
        ).get(id=gs.id)

        return Response(GroupScheduleSerializer(gs).data)


    @action(detail=False, methods=['post'], url_path='create-draft')
    def create_draft(self, request):
        """Создаёт черновик расписания для зачисления"""
        import traceback

        try:
            serializer = GroupScheduleCreateSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)

            enrollment_id = serializer.validated_data['enrollment_id']

            try:
                enrollment = Enrollment.objects.select_related('module').get(id=enrollment_id)
            except Enrollment.DoesNotExist:
                return Response(
                    {'error': f'Enrollment с id={enrollment_id} не найден'},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Проверяем существующее расписание
            existing = GroupSchedule.objects.filter(
                enrollment=enrollment,
                status__in=[ScheduleStatus.DRAFT, ScheduleStatus.ACTIVE]
            ).first()

            if existing:
                existing = GroupSchedule.objects.prefetch_related(
                    'items__stage', 'items__section', 'items__instructor'
                ).get(id=existing.id)
                return Response(
                    GroupScheduleSerializer(existing).data,
                    status=status.HTTP_200_OK
                )

            # Проверяем stages
            stages = enrollment.module.stages.prefetch_related('sections').all().order_by('order')
            if not stages.exists():
                return Response(
                    {'error': 'У модуля нет этапов (stages).'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 🔹 Формируем значение для group_code из данных зачисления
            safe_group_code = str(enrollment.group) if enrollment.group else f"Группа_{enrollment.id}"

            # Создаём черновик
            with transaction.atomic():
                gs = GroupSchedule.objects.create(
                    enrollment=enrollment,
                    group_code=safe_group_code,  # 🔹 ТЕПЕРЬ ЭТО РАБОТАЕТ
                    status=ScheduleStatus.DRAFT,
                    version=1
                )

                # 🔹 Копируем структуру модуля с ПОСЛЕДОВАТЕЛЬНОЙ нумерацией (начиная с 1)
                items = []
                item_order = 1

                for stage in stages:
                    sections = stage.sections.all().order_by('order')
                    if not sections.exists():
                        continue

                    for section in sections:
                        items.append(ScheduleItem(
                            group_schedule=gs,
                            stage=stage,
                            section=section,
                            order=item_order  # 🔹 Последовательный номер: 1, 2, 3...
                        ))
                        item_order += 1

                if not items:
                    gs.delete()
                    return Response(
                        {'error': 'У модуля нет разделов (sections).'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                ScheduleItem.objects.bulk_create(items, batch_size=500)

            # Перезагружаем с связями для отдачи
            gs = GroupSchedule.objects.prefetch_related(
                'items__stage', 'items__section', 'items__instructor'
            ).get(id=gs.id)

            return Response(
                GroupScheduleSerializer(gs).data,
                status=status.HTTP_201_CREATED
            )

        except Exception as e:
            print("=" * 50)
            print("❌ ОШИБКА В create_draft:")
            print(traceback.format_exc())
            print("=" * 50)
            return Response(
                {'error': f'Внутренняя ошибка: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    @action(detail=True, methods=['post'], url_path='update-items')
    def update_items(self, request, pk=None):
        """Массовое обновление элементов расписания"""
        gs = self.get_object()

        if gs.status == ScheduleStatus.FINALIZED:
            return Response(
                {'error': 'Расписание уже зафиксировано'},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = ScheduleItemBulkUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        items_data = serializer.validated_data['items']
        updated_ids = []

        with transaction.atomic():
            for item_data in items_data:
                item_id = item_data.pop('id')
                try:
                    item = ScheduleItem.objects.get(id=item_id, group_schedule=gs)
                except ScheduleItem.DoesNotExist:
                    continue

                if item.is_locked:
                    continue  # Пропускаем заблокированные

                # Обработка instructor_id → instructor
                if 'instructor_id' in item_data:
                    instructor_id = item_data.pop('instructor_id')
                    if instructor_id:
                        item.instructor_id = instructor_id
                    else:
                        item.instructor = None

                # Обновляем остальные поля
                for key, value in item_data.items():
                    setattr(item, key, value)

                item.save()
                updated_ids.append(item.id)

        # Возвращаем обновлённое расписание
        gs = GroupSchedule.objects.prefetch_related(
            'items__stage', 'items__section', 'items__instructor'
        ).get(id=gs.id)

        return Response(
            GroupScheduleSerializer(gs).data,
            status=status.HTTP_200_OK
        )

    @action(detail=True, methods=['post'], url_path='finalize')
    def finalize(self, request, pk=None):
        """Утверждение расписания"""
        gs = self.get_object()

        # Проверяем, что все обязательные поля заполнены
        incomplete = gs.items.filter(
            date__isnull=True
        ).exists()

        if incomplete:
            return Response(
                {'error': 'Не у всех элементов заполнена дата'},
                status=status.HTTP_400_BAD_REQUEST
            )

        gs.status = ScheduleStatus.FINALIZED
        gs.save()

        return Response(
            GroupScheduleSerializer(gs).data,
            status=status.HTTP_200_OK
        )

    @action(detail=True, methods=['post'], url_path='reset-to-draft')
    def reset_to_draft(self, request, pk=None):
        """Сброс в черновик"""
        gs = self.get_object()
        gs.status = ScheduleStatus.DRAFT
        gs.save()
        return Response(GroupScheduleSerializer(gs).data)

    @action(detail=False, methods=['get'], url_path='by-enrollment/(?P<enrollment_id>[^/.]+)')
    def by_enrollment(self, request, enrollment_id=None):
        """Получить расписание по enrollment_id"""
        gs = GroupSchedule.objects.filter(enrollment_id=enrollment_id).first()
        if not gs:
            return Response({'detail': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        gs = GroupSchedule.objects.prefetch_related(
            'items__stage', 'items__section', 'items__instructor'
        ).get(id=gs.id)
        return Response(GroupScheduleSerializer(gs).data)

    @action(detail=False, methods=['get'], url_path='staff-list')
    def staff_list(self, request):
        """Список преподавателей для выбора"""
        try:
            staff = Staff.objects.filter(is_active=True).order_by('name')

            data = [{
                'id': s.id,
                'full_name': s.name or f"Сотрудник #{s.id}",
                'position': getattr(s, 'position', '')
            } for s in staff]

            return Response(data)

        except Exception as e:
            import traceback
            print("❌ ОШИБКА В staff_list:")
            print(traceback.format_exc())
            return Response({'error': str(e)}, status=500)



    @action(detail=False, methods=['post'], url_path='export-excel')
    def export_excel(self, request):
        """Экспорт расписания в Excel"""
        enrollment_id = request.data.get('enrollment_id')
        if not enrollment_id:
            return HttpResponse('❌ enrollment_id is required', status=400)

        try:
            enrollment = Enrollment.objects.select_related('module__course').get(id=enrollment_id)
        except Enrollment.DoesNotExist:
            return HttpResponse('❌ Enrollment not found', status=404)

        # 🔹 1. Находим или создаём черновик расписания
        gs, created = GroupSchedule.objects.get_or_create(
            enrollment=enrollment,
            defaults={'status': 'draft', 'version': 1}
        )

        # 🔹 2. Если элементов нет — генерируем скелет из модуля
        if gs.items.count() == 0:
            items_to_create = []
            for stage in enrollment.module.stages.prefetch_related('sections').all().order_by('order'):
                for section in stage.sections.all().order_by('order'):
                    items_to_create.append(ScheduleItem(
                        group_schedule=gs,
                        stage=stage,
                        section=section,
                        order=section.order
                    ))
            ScheduleItem.objects.bulk_create(items_to_create, batch_size=500)

        # 🔹 3. Загружаем элементы с данными
        items = ScheduleItem.objects.select_related(
            'section', 'stage', 'instructor', 'group_schedule__enrollment'
        ).filter(group_schedule=gs).order_by('stage__order', 'section__order')

        module = enrollment.module
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Расписание"

        # Стили
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
        bold_font = Font(bold=True)
        stage_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        # 🔹 Шапка (данные берём из GroupSchedule)
        ws.merge_cells('C1:E1')
        ws[
            'C1'] = f"Программа подготовки специалистов согласно перечню специалистов авиационного персонала гражданской авиации «Периодическая наземная подготовка членов летных экипажей»"
        ws['C1'].alignment = align_left
        ws['C1'].font = Font(bold=True, size=11)

        ws.merge_cells('C2:E2')
        ws[
            'C2'] = f"Модуль 3. Ежегодная аварийно-спасательная подготовка и подготовка по управлению ресурсами экипажа ВС Boeing -737-800 (Next Generation - NG)" if module else ""
        ws['C2'].alignment = align_left
        ws['C2'].font = Font(bold=True, size=11)

        ws.merge_cells('C3:E3')
        ws['C3'] = f"Группа: {enrollment.group}"
        ws['C3'].alignment = align_left
        ws['C3'].font = Font(bold=True, size=11)

        ws.merge_cells('C4:E4')
        ws['C4'] = f"Дата начала: {gs.start_date.strftime('%d.%m.%Y') if gs.start_date else ''}"
        ws['C4'].alignment = align_left

        ws['A5'] = "Куратор"
        ws['A5'].font = bold_font
        ws.merge_cells('B5:E5')
        ws['B5'] = gs.curator.name if gs.curator and hasattr(gs.curator, 'name') else ""
        ws['G4'] = "УТВЕРЖДАЮ:"
        ws['H4'].font = bold_font
        ws.merge_cells('G5:H5')
        ws['G5'] = "Директор(Заместитель директора)"
        ws['G5'].font = bold_font
        ws.merge_cells('G6:H6')
        ws['G6'] = gs.director.name if gs.director and hasattr(gs.director, 'name') else ""

        # 🔹 Заголовки таблицы
        ws.append([''] * 8)
        headers = ['№', 'Дата', 'Раздел', 'Подраздел/Сессия', 'Продолжительность', 'Почасовое расписание',
                   'Преподаватель',
                   'Место проведения']
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = bold_font
            cell.alignment = align_center
            cell.border = thin_border

        row_num = 1
        current_stage_id = None

        for item in items:
            # 🔹 Заголовок этапа (выводим только при смене) - БЕЗ stage и скобок
            if item.stage.id != current_stage_id:
                ws.append([''] * 8)
                r = ws.max_row
                ws.merge_cells(f'A{r}:H{r}')
                # 🔹 Убираем скобки и "stage"
                stage_title = item.stage.title.replace('(', '').replace(')', '').replace('stage', '').strip()
                ws[f'A{r}'] = f"{stage_title}"
                ws[f'A{r}'].font = bold_font
                ws[f'A{r}'].fill = stage_fill
                current_stage_id = item.stage.id

            # 🔹 Генерация времени на лету
            start = item.start_time or "09:00"
            times = get_schedule_times(item.effective_detail, start_time=start)

            # 🔹 Подраздел (если есть связь)
            sub_title = ""
            if hasattr(item.section, 'subsection') and item.section.subsection:
                sub_title = item.section.subsection.title
            elif hasattr(item.section, 'subsections') and item.section.subsections.exists():
                sub_title = ", ".join([s.title for s in item.section.subsections.all()[:2]])

            ws.append([
                row_num,
                item.date.strftime('%d.%m.%Y') if item.date else '',
                item.section.title,
                sub_title,
                float(item.section.duration_hours) if item.section.duration_hours else 0,
                times,
                item.instructor.name if item.instructor and hasattr(item.instructor, 'name') else '',
                item.get_location_display() if item.location else ''
            ])

            r = ws.max_row
            for col_idx, cell in enumerate(ws[r], 1):
                cell.border = thin_border
                cell.alignment = align_left if col_idx in [3, 4, 6] else align_center
            row_num += 1

        # Ширины колонок
        for col, width in zip('ABCDEFGH', [5, 12, 30, 25, 10, 45, 25, 25]):
            ws.column_dimensions[col].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="schedule_{module.code}_{enrollment.group}.xlsx"'
        wb.save(response)
        return response
# =============================================================================
# РАСПИСАНИЕ В ПАПКУ ГРУППЫ
# =============================================================================


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_schedule_to_folder(request):
    """POST /api/schedules/export-to-folder/ — Сохраняет расписание в папку группы"""

    data = request.data
    group_code = data.get('group')

    if not group_code:
        return Response({'error': 'Не указан номер группы'}, status=400)

    try:
        # 1. Генерируем Excel (байты)
        excel_bytes = generate_schedule_excel(
            group=data.get('group'),
            module_code=data.get('module_code'),
            module=data.get('module'),
            course=data.get('course'),
            start_date=data.get('start_date'),
            curator=data.get('curator'),
            director=data.get('director'),
            schedule=data.get('schedule', [])
        )

        # 2. Создаём/проверяем папку группы
        folder_path = ensure_group_folder(group_code)

        # 3. Формируем имя файла
        filename = f"Расписание_{group_code}.xlsx"
        file_path = os.path.join(folder_path, filename)

        # 4. Сохраняем файл ПРЯМО В ПАПКУ ГРУППЫ
        with open(file_path, 'wb') as f:
            f.write(excel_bytes)

        # 5. Возвращаем относительный путь
        relative_path = os.path.join('groups', group_code, filename)

        return Response({
            'success': True,
            'message': 'Расписание сохранено в папку группы',
            'file_path': relative_path
        }, status=200)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': f'Ошибка сохранения: {str(e)}'}, status=500)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_staff_list(request):
    """GET /api/staff/ — возвращает список персонала для выбора"""
    # Фильтруем активных сотрудников с должностями
    staff = Staff.objects.filter(is_active=True).exclude(
        name__isnull=True
    ).exclude(name='').values('id', 'name', 'position', 'fptitle', 'tptitle').order_by('name')

    return Response(list(staff))

# ====================================================================================================================
# Экспорт расписания EXEL
# ====================================================================================================================

# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def export_schedule_excel(request):
#     enrollment_id = request.data.get('enrollment_id')
#     if not enrollment_id:
#         return HttpResponse('❌ enrollment_id is required', status=400)
#
#     try:
#         enrollment = Enrollment.objects.select_related('module__course').get(id=enrollment_id)
#     except Enrollment.DoesNotExist:
#         return HttpResponse('❌ Enrollment not found', status=404)
#
#     # 🔹 1. Находим или создаём черновик расписания
#     gs, created = GroupSchedule.objects.get_or_create(
#         enrollment=enrollment,
#         defaults={'status': 'draft', 'version': 1}
#     )
#
#     # 🔹 2. Если элементов нет — генерируем скелет из модуля
#     if gs.items.count() == 0:
#         items_to_create = []
#         for stage in enrollment.module.stages.prefetch_related('sections').all().order_by('order'):
#             for section in stage.sections.all().order_by('order'):
#                 items_to_create.append(ScheduleItem(
#                     group_schedule=gs,
#                     stage=stage,
#                     section=section,
#                     order=section.order
#                 ))
#         ScheduleItem.objects.bulk_create(items_to_create, batch_size=500)
#
#     # 🔹 3. Загружаем элементы с данными
#     items = ScheduleItem.objects.select_related(
#         'section', 'stage', 'instructor', 'group_schedule__enrollment'
#     ).filter(group_schedule=gs).order_by('stage__order', 'section__order')
#
#     module = enrollment.module
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Расписание"
#
#     # Стили
#     thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
#                          bottom=Side(style='thin'))
#     align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
#     align_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
#     bold_font = Font(bold=True)
#     stage_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
#
#     #  Шапка (данные берём из GroupSchedule)
#     ws.merge_cells('C1:E1');
#     ws['C1'] = f"Программа подготовки (course)\n{module.course.title}";
#     ws['C1'].alignment = align_left
#     ws.merge_cells('C2:E2');
#     ws['C2'] = f"Модуль (module)\n{module.title}";
#     ws['C2'].alignment = align_left
#     ws.merge_cells('C3:E3');
#     ws['C3'] = f"Группа (group)\n{enrollment.group}";
#     ws['C3'].alignment = align_left
#     ws.merge_cells('C4:E4');
#     ws['C4'] = f"Дата начала\n{gs.start_date.strftime('%d.%m.%Y') if gs.start_date else ''}";
#     ws['C4'].alignment = align_left
#
#     ws['A5'] = "Куратор";
#     ws['A5'].font = bold_font
#     ws.merge_cells('B5:E5');
#     ws['B5'] = gs.curator.get_full_name() if gs.curator else ""
#     ws['F5'] = "УТВЕРЖДАЮ:";
#     ws['F5'].font = bold_font
#     ws.merge_cells('G5:H5');
#     ws['G5'] = "Директор"
#     ws['G6'] = gs.director.get_full_name() if gs.director else ""
#
#     # 🔹 Заголовки таблицы
#     ws.append([''] * 8)
#     headers = ['№', 'Дата', 'Раздел', 'Подраздел/Сессия', 'Продолжительность', 'Почасовое расписание', 'Преподаватель',
#                'Место проведения']
#     ws.append(headers)
#     for cell in ws[ws.max_row]:
#         cell.font = bold_font
#         cell.alignment = align_center
#         cell.border = thin_border
#
#     row_num = 1
#     current_stage_id = None
#
#     for item in items:
#         # 🔹 Заголовок этапа (выводим только при смене)
#         if item.stage.id != current_stage_id:
#             ws.append([''] * 8)
#             r = ws.max_row
#             ws.merge_cells(f'A{r}:H{r}')
#             ws[f'A{r}'] = f"({item.stage.title}) stage"
#             ws[f'A{r}'].font = bold_font
#             ws[f'A{r}'].fill = stage_fill
#             current_stage_id = item.stage.id
#
#         # 🔹 Генерация времени на лету
#         start = item.start_time or "09:00"
#         times = get_schedule_times(item.effective_detail, start_time=start)
#
#         # 🔹 Подраздел (если есть связь)
#         sub_title = ""
#         try:
#             if hasattr(item.section, 'subsection') and item.section.subsection:
#                 sub_title = item.section.subsection.title
#             elif hasattr(item.section, 'subsections') and item.section.subsections.exists():
#                 sub_title = ", ".join([s.title for s in item.section.subsections.all()[:2]])
#         except Exception:
#             sub_title = ""  # Если нет связи, оставляем пустым
#
#         ws.append([
#             row_num,
#             item.date.strftime('%d.%m.%Y') if item.date else '',
#             item.section.title,
#             sub_title,
#             float(item.section.duration_hours) if item.section.duration_hours else 0,
#             times,
#             item.instructor.get_full_name() if item.instructor else '',
#             item.get_location_display() if item.location else ''
#         ])
#
#         r = ws.max_row
#         for col_idx, cell in enumerate(ws[r], 1):
#             cell.border = thin_border
#             cell.alignment = align_left if col_idx in [3, 4, 6] else align_center
#         row_num += 1
#
#     # Ширины колонок
#     for col, width in zip('ABCDEFGH', [5, 12, 30, 25, 10, 45, 25, 25]):
#         ws.column_dimensions[col].width = width
#
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = f'attachment; filename="schedule_{module.code}_{enrollment.group}.xlsx"'
#     wb.save(response)
#     return response
# ====================================================================================================================
# Создание, обновление расписания
# ====================================================================================================================


# ====================================================================================================================
# Сохранение черновика расписания
# ====================================================================================================================



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def module_list(request):
    """GET /api/modules/ — список всех модулей для выпадающего списка"""
    from api.models import Module
    modules = Module.objects.select_related('course').order_by('code')
    data = [
        {
            'id': m.id,
            'code': m.code,
            'title': m.title
        }
        for m in modules
    ]
    return Response(data)

# =============================================================================
# ПАПКИ ГРУПП
# =============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_group_folders(request):
    """GET /api/groups/folders/ — Список папок групп и их содержимого"""

    groups_path = os.path.join(settings.MEDIA_ROOT, 'groups')

    # Если папки ещё нет — создаём (чтобы не было ошибки)
    os.makedirs(groups_path, exist_ok=True)

    folders = []

    # 🔹 Перебираем папки групп
    for item in os.listdir(groups_path):
        item_path = os.path.join(groups_path, item)

        if os.path.isdir(item_path):
            # 🔹 Список файлов в папке группы
            files = []
            for f in os.listdir(item_path):
                f_path = os.path.join(item_path, f)
                if os.path.isfile(f_path):
                    files.append({
                        'name': f,
                        'size': os.path.getsize(f_path),
                        'modified': os.path.getmtime(f_path),
                        'url': f'/media/groups/{item}/{f}'  # относительный URL
                    })

            folders.append({
                'name': item,
                'path': os.path.join('groups', item),
                'created': os.path.getctime(item_path),
                'files_count': len(files),
                'files': files  # опционально: можно не отдавать сразу
            })

    # 🔹 Сортируем по имени (или по дате создания)
    folders.sort(key=lambda x: x['name'], reverse=True)

    return Response({'folders': folders}, status=200)