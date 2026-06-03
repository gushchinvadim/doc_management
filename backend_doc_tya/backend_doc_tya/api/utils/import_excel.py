# api/utils/import_excel.py
import re
import logging
from datetime import datetime, timedelta
from api.models import Course, Module, Stage, Section, Staff, Student, SubSection
import openpyxl
from django.contrib.auth.models import User
from django.db.utils import IntegrityError
import openpyxl
import logging
from decimal import Decimal






logger = logging.getLogger(__name__)
EXCEL_EPOCH = datetime(1899, 12, 30)

# =============================================================================
# УТИЛИТЫ
# =============================================================================
def parse_excel_date(val):
    if val is None: return None
    if isinstance(val, datetime): return val.date()
    if isinstance(val, (int, float)):
        return (EXCEL_EPOCH + timedelta(days=val)).date()
    for fmt in ('%m/%d/%y', '%d/%m/%y', '%Y-%m-%d', '%d.%m.%Y'):
        try:
            return datetime.strptime(str(val).strip(), fmt).date()
        except ValueError:
            continue
    return None


def _clean(val):
    """Универсальная очистка строк (аналог pandas .str)"""
    if val is None: return ''
    s = str(val).strip()
    if s.lower() in ('nan', 'none', ''): return ''
    s = s.replace('Пограмма', 'Программа')
    s = re.sub(r'[\n\r\t]+', ' ', s).strip()
    return s


def _to_float(val):
    try:
        return float(val) if val is not None else 0.0
    except (ValueError, TypeError):
        return 0.0


def _to_int(val):
    try:
        return int(float(val)) if val is not None else 0
    except (ValueError, TypeError):
        return 0


# =============================================================================
# COURSE
# =============================================================================

def import_courses_excel(file_obj):
    # 🔹 Вспомогательные функции
    def _clean(val):
        if val is None: return ''
        s = str(val).strip()
        return '' if s.lower() in ('nan', 'none', 'null', 'нет', '') else s

    def _to_float(val):
        if val is None: return None
        try: return float(str(val).replace(',', '.'))
        except (ValueError, TypeError): return None

    def _to_int(val):
        if val is None: return None
        try: return int(float(str(val).replace(',', '.')))
        except (ValueError, TypeError): return None

    wb = openpyxl.load_workbook(file_obj, data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(min_row=1, values_only=True))
    if not rows:
        return 0, 0, 0, ["Файл пуст"]

    headers = [str(h).strip().upper() for h in rows[0]]
    data_rows = rows[1:]

    created = updated = errors = 0
    error_details = []

    course_cache = {}
    module_cache = {}
    stage_cache = {}
    section_cache = {}
    subsection_cache = {}
    sub_counters = {}

    # 🔹 Маппинг значений из колонки DETAILS в ключи модели
    DETAIL_MAP = {
        'самостоятельная подготовка в сдо': 'sdo',
        'сдо': 'sdo',
        '1 час - брифинг (лекция)': 'sim',
        'тренажерная подготовка': 'sim',
        'sim': 'sim',
    }
    VALID_DETAILS = {k for k, _ in Section.DETAIL_CHOICES}

    for i, row in enumerate(data_rows, start=2):
        if not any(v is not None for v in row):
            continue
        try:
            row_dict = {headers[idx]: val for idx, val in enumerate(row) if idx < len(headers)}

            comp_code = _clean(row_dict.get('COMPANY_CODE'))
            if not comp_code: continue

            # 1️⃣ COURSE
            if comp_code not in course_cache:
                prog_id = _clean(row_dict.get('PROG_ID', ''))[:10]
                course_defaults = {'title': _clean(row_dict.get('COURSE')), 'prog_id': prog_id}
                course, c_created = Course.objects.get_or_create(company_code=comp_code, defaults=course_defaults)
                if c_created: created += 1
                else: updated += 1
                course_cache[comp_code] = course
            course = course_cache[comp_code]

            # 2️⃣ MODULE
            mod_id = _clean(row_dict.get('MOD_ID'))
            if not mod_id: continue

            att_raw = row_dict.get('ATTACHMENT_NUMBER')
            att_num = int(float(att_raw)) if att_raw is not None and str(att_raw).strip() else 0

            mod_defaults = {
                'title': _clean(row_dict.get('MODULE')),
                'duration': _to_float(row_dict.get('DURATION')),
                'code': _clean(row_dict.get('CODE', '')),
                'attachment_number': att_num,
            }
            module, m_created = Module.objects.update_or_create(course=course, mod_id=mod_id, defaults=mod_defaults)
            if m_created: created += 1
            else: updated += 1
            module_cache[(course.id, mod_id)] = module
            module = module_cache[(course.id, mod_id)]

            # 3️ STAGE
            stage_title = _clean(row_dict.get('STAGE'))
            if not stage_title: continue

            stage_key = (module.id, stage_title)
            if stage_key not in stage_cache:
                current_order = Stage.objects.filter(module=module).count()
                stage, s_created = Stage.objects.get_or_create(
                    module=module, title=stage_title, defaults={'order': current_order, 'description': ''}
                )
                if s_created: created += 1
                stage_cache[stage_key] = stage
            stage = stage_cache[stage_key]

            # 4️⃣ SECTION
            sec_title = _clean(row_dict.get('SECTION'))
            if not sec_title: continue

            sub_sec_title = _clean(row_dict.get('SUB_SECTION'))

            # 🔹 Определяем: есть ли реальный подраздел?
            # Подраздел реальный, если:
            # 1. Он не пустой
            # 2. Он НЕ совпадает с названием раздела (иначе это "ложный" подраздел)
            has_real_subsection = (
                    sub_sec_title
                    and sub_sec_title != sec_title
                    and sub_sec_title.lower() not in ('nan', 'none', 'null', 'нет', '')
            )

            sec_order_raw = row_dict.get('ORDER')
            sec_order = _to_int(sec_order_raw) if sec_order_raw is not None else Section.objects.filter(
                stage=stage).count()

            grade_raw = _clean(row_dict.get('GRADE_TYPE', 'none')).lower()
            sec_grade = grade_raw if grade_raw in ['numeric', 'binary', 'none'] else 'none'

            min_raw = row_dict.get('MIN_SCORE')
            sec_min = 1 if sec_grade == 'binary' else (
                _to_int(min_raw) if min_raw is not None and _clean(str(min_raw)) else None)

            # 🔹 Обработка DETAILS для раздела
            detail_raw = _clean(row_dict.get('DETAILS', '')).lower()
            detail_val = DETAIL_MAP.get(detail_raw, detail_raw)
            if detail_val not in VALID_DETAILS:
                detail_val = 'sdo'

            # 🔹 ВАЖНО: если есть реальный подраздел — НЕ перезаписываем detail раздела
            # (он может быть разным для разных подразделов одного раздела)
            sec_defaults = {
                'duration_hours': _to_float(row_dict.get('DURATION_HOURS')),
                'grade_type': sec_grade,
                'min_score': sec_min,
                'order': sec_order,
                # Если подраздела нет — сохраняем detail в раздел.
                # Если есть — оставляем section.detail как есть (не трогаем).
            }

            # Устанавливаем section.detail только если подраздела НЕТ
            if not has_real_subsection:
                sec_defaults['detail'] = detail_val

            section, sec_created = Section.objects.update_or_create(
                stage=stage, title=sec_title, defaults=sec_defaults
            )
            if sec_created:
                created += 1
            else:
                updated += 1
            section_cache[(stage.id, sec_title)] = section
            section = section_cache[(stage.id, sec_title)]

            # 5️⃣ SUB_SECTION
            if has_real_subsection:
                sub_key = (section.id, sub_sec_title)
                if sub_key not in subsection_cache:
                    if section.id not in sub_counters:
                        sub_counters[section.id] = 0
                    sub_counters[section.id] += 1

                    # 🔹 ВАЖНО: detail из Excel идёт именно в подраздел!
                    sub_defaults = {
                        'duration_hours': _to_float(row_dict.get('DURATION_HOURS')),
                        'order': sub_counters[section.id],
                        'detail': detail_val,  # ← ВОТ ЭТО ГЛАВНОЕ ИЗМЕНЕНИЕ
                    }

                    sub_sec, sub_created = SubSection.objects.update_or_create(
                        section=section, title=sub_sec_title, defaults=sub_defaults
                    )
                    if sub_created:
                        created += 1
                    else:
                        updated += 1
                    subsection_cache[sub_key] = sub_sec

            # 5️⃣ SUB_SECTION
            sub_sec_title = _clean(row_dict.get('SUB_SECTION'))
            if sub_sec_title:
                sub_key = (section.id, sub_sec_title)
                if sub_key not in subsection_cache:
                    if section.id not in sub_counters:
                        sub_counters[section.id] = 0
                    sub_counters[section.id] += 1

                    # 🔹 НОВОЕ: Обработка колонки DETAILS для подраздела
                    sub_detail_raw = _clean(row_dict.get('DETAILS', '')).lower()
                    sub_detail_val = DETAIL_MAP.get(sub_detail_raw, sub_detail_raw)

                    # Если в Excel для подраздела detail не валиден или пуст, берем от родительского раздела
                    if sub_detail_val not in VALID_DETAILS:
                        sub_detail_val = section.detail or 'sdo'

                    sub_defaults = {
                        'duration_hours': _to_float(row_dict.get('DURATION_HOURS')),
                        'order': sub_counters[section.id],
                        'detail': sub_detail_val,  # 🔹 ДОБАВЛЕНО: сохраняем detail в подраздел!
                    }

                    sub_sec, sub_created = SubSection.objects.update_or_create(
                        section=section, title=sub_sec_title, defaults=sub_defaults
                    )
                    if sub_created:
                        created += 1
                    else:
                        updated += 1
                    subsection_cache[sub_key] = sub_sec
        except Exception as e:
            errors += 1
            err_msg = f"Строка {i}: {e}"
            error_details.append(err_msg)
            logger.warning(f"[Course Import] {err_msg} | Row: {row}")

    return created, updated, errors, error_details

# =============================================================================
# STAFF
# =============================================================================


def _parse_bool(val):
    """Безопасное преобразование значения Excel в boolean"""
    if val is None or str(val).strip() in ('', '0', 'False', 'false', 'Нет', 'нет', 'No', 'no', 'N', 'n'):
        return False
    return True


def import_staff_excel(file_obj):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))

    created = updated = errors = 0
    error_details = []

    for i, row in enumerate(rows, start=2):
        # Пропускаем полностью пустые строки
        if not any(row):
            continue

        try:
            # ✅ Ожидаем минимум 6 колонок
            if len(row) < 6:
                raise ValueError(f"Недостаточно колонок (ожидается 6, получено {len(row)})")

            # Берём только первые 6 значений, чтобы избежать IndexError при пустых хвостах
            name, rauts_id, position, is_active_raw, fptitle_raw, tptitle_raw = row[:6]

            username = str(rauts_id).strip() if rauts_id else ''
            if not username:
                raise ValueError("Пустой rauts_id")

            # 🔹 Парсим булевы поля (пустая ячейка = False, "1"/"Да"/"True" = True)
            is_active = _parse_bool(is_active_raw)
            fptitle = _parse_bool(fptitle_raw)
            tptitle = _parse_bool(tptitle_raw)

            # 👤 1. Находим или создаём User
            user, user_created = User.objects.get_or_create(username=username)
            if user_created:
                user.set_unusable_password()  # 🔒 Вход запрещён до установки пароля
                user.is_active = is_active
                user.save()
            elif user.is_active != is_active:
                user.is_active = is_active
                user.save(update_fields=['is_active'])

            # 📋 2. Создаём/обновляем Staff
            _, is_created = Staff.objects.update_or_create(
                rauts_id=username,
                defaults={
                    'user': user,
                    'name': str(name).strip() if name else '',
                    'position': str(position).strip() if position else '',
                    'is_active': is_active,
                    'fptitle': fptitle,  # ← новое поле
                    'tptitle': tptitle,  # ← новое поле
                }
            )
            if is_created:
                created += 1
            else:
                updated += 1

        except IntegrityError as e:
            errors += 1
            err_msg = f"Строка {i}: Конфликт уникального rauts_id '{username}'"
            error_details.append(err_msg)
            logger.warning(f"[Staff Import] {err_msg}")
        except Exception as e:
            errors += 1
            err_msg = f"Строка {i}: {e}"
            error_details.append(err_msg)
            logger.warning(f"[Staff Import] {err_msg} | Data: {row}")

    return created, updated, errors, error_details

# =============================================================================
# STUDENT
# =============================================================================
def import_students_excel(file_obj):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    created = updated = errors = 0
    error_details = []

    for i, row in enumerate(rows, start=2):
        if not any(v is not None for v in row):
            continue
        try:
            # 🔹 Точный порядок колонок из students.xlsx (14 полей)
            if len(row) < 14:
                raise ValueError(f"Недостаточно колонок (ожидается 14, получено {len(row)})")

            surname = row[0]
            name = row[1]
            patronymic = row[2]
            sex = row[3]
            dob_raw = row[4]
            snils = row[5]
            surname_latin = row[6]
            name_latin = row[7]
            employee_id = row[8]
            dcat_id_raw = row[9]
            citizenship_code_raw = row[10]
            email_raw = row[11]
            is_active_raw = row[12]
            aircraft_type_raw = row[13]

            # 🔹 Валидация обязательных полей
            if not snils or not surname or not name:
                raise ValueError("Отсутствует СНИЛС, Фамилия или Имя")

            # 🔹 Очистка данных
            snils = str(snils).replace(' ', '').replace('-', '')
            surname = str(surname).strip()
            name = str(name).strip()
            patronymic = str(patronymic).strip() if patronymic else ''
            surname_latin = str(surname_latin).strip() if surname_latin else ''
            name_latin = str(name_latin).strip() if name_latin else ''
            employee_id = str(employee_id).strip() if employee_id else None
            citizenship_code = str(citizenship_code_raw).strip() if citizenship_code_raw else '643'
            email = str(email_raw).strip() if email_raw else ''
            if email.lower() in ('nan', 'none', 'null', 'нет', ''):
                email = ''

            # 🔹 Пол
            sex_map = {'Муж': 'Male', 'Жен': 'Female', 'М': 'Male', 'Ж': 'Female', '1': 'Male', '0': 'Female'}
            parsed_sex = sex_map.get(str(sex).strip(), 'Male') if sex else 'Male'

            # 🔹 Дата рождения
            dob_date = parse_excel_date(dob_raw)
            if not dob_date:
                raise ValueError(f"Неверный формат даты: {dob_raw}")

            # 🔹 Активность
            is_active = str(is_active_raw).strip() in ['1', 'True', 'true', 'Да', 'да', '1.0', True]

            # 🔹 dcat_id (только int 1, 2, 3)
            dcat_id = None
            if dcat_id_raw is not None and str(dcat_id_raw).strip():
                try:
                    val = int(str(dcat_id_raw).strip())
                    if val in [1, 2, 3]:
                        dcat_id = val
                except ValueError:
                    pass

            # 🔹 Тип ВС (строго по choices модели)
            valid_ac = ['B737NG', 'B737CL', 'B737CL+B737NG']
            ac_parsed = None
            if aircraft_type_raw:
                ac_clean = str(aircraft_type_raw).strip()
                if ac_clean in valid_ac:
                    ac_parsed = ac_clean

            # 🔹 Создание/обновление по СНИЛС (уникальный ключ модели)
            _, is_created = Student.objects.update_or_create(
                snils=snils,
                defaults={
                    'surname': surname,
                    'name': name,
                    'patronymic': patronymic,
                    'sex': parsed_sex,
                    'date_of_birth': dob_date,
                    'employee_id': employee_id,
                    'is_active': is_active,
                    'aircraft_type': ac_parsed,
                    'surname_latin': surname_latin,
                    'name_latin': name_latin,
                    'dcat_id': dcat_id,
                    'citizenship_code': citizenship_code,
                    'email': email,
                }
            )
            if is_created:
                created += 1
            else:
                updated += 1

        except IntegrityError as e:
            errors += 1
            err_msg = f"Строка {i}: Ошибка уникальности (СНИЛС/employee_id)"
            error_details.append(err_msg)
            logger.warning(f"[Student Import] {err_msg} | {e}")
        except Exception as e:
            errors += 1
            err_msg = f"Строка {i}: {e}"
            error_details.append(err_msg)
            logger.warning(f"[Student Import] {err_msg} | Data: {row}")

    return created, updated, errors, error_details
# =============================================================================
# REGISTRY
# =============================================================================
IMPORT_REGISTRY = {
    'course': import_courses_excel,
    'staff': import_staff_excel,
    'student': import_students_excel,
}