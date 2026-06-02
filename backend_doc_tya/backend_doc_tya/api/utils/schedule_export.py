# api/utils/schedule_export.py
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime


def generate_schedule_excel(group, module_code, module, course, start_date,
                            curator, director, schedule):
    """
    Генерирует Excel-файл расписания для группы.
    Возвращает байты файла.
    """
    # Создаём новую книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Расписание"

    # 🔹 Стили
    header_font = Font(bold=True, size=12, name='Times New Roman')
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 🔹 Заголовок документа
    ws.merge_cells('A1:H1')
    ws['A1'] = f"РАСПИСАНИЕ ЗАНЯТИЙ"
    ws['A1'].font = Font(bold=True, size=16, name='Times New Roman')
    ws['A1'].alignment = center_align

    # Группа и модуль
    ws.merge_cells('A2:H2')
    ws['A2'] = f"Группа: {group} | Модуль: {module_code}"
    ws['A2'].font = Font(bold=True, size=11, name='Times New Roman')
    ws['A2'].alignment = center_align

    # Курс
    ws.merge_cells('A3:H3')
    ws['A3'] = f"Курс: {course}" if course else ""
    ws['A3'].font = Font(size=11, name='Times New Roman')
    ws['A3'].alignment = center_align

    # Дата начала
    ws.merge_cells('A4:H4')
    ws['A4'] = f"Дата начала: {start_date}" if start_date else ""
    ws['A4'].font = Font(size=11, name='Times New Roman')
    ws['A4'].alignment = center_align

    # Куратор и директор
    ws.merge_cells('A5:H5')
    ws['A5'] = f"Куратор: {curator} | Утверждающий: {director}"
    ws['A5'].font = Font(size=11, name='Times New Roman')
    ws['A5'].alignment = center_align

    # 🔹 Заголовки таблицы
    headers = [
        '№ п/п', 'Дата', 'День недели', 'Время', 'Дисциплина/Этап',
        'Тема занятия', 'Инструктор', 'Место проведения', 'Часы'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    # 🔹 Данные расписания
    current_stage = ""
    row_num = 8

    for idx, item in enumerate(schedule, 1):
        # Если есть этап и он изменился — добавляем строку с названием этапа
        if item.get('stage') and item['stage'] != current_stage:
            ws.merge_cells(f'A{row_num}:H{row_num}')
            stage_cell = ws.cell(row=row_num, column=1, value=f"ЭТАП: {item['stage']}")
            stage_cell.font = Font(bold=True, size=11, name='Times New Roman')
            stage_cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            stage_cell.alignment = center_align
            stage_cell.border = thin_border
            row_num += 1
            current_stage = item['stage']

        # Строка с данными
        ws.cell(row=row_num, column=1, value=idx).alignment = center_align
        ws.cell(row=row_num, column=2, value=item.get('date', '')).alignment = center_align
        ws.cell(row=row_num, column=3, value=item.get('weekday', '')).alignment = center_align
        ws.cell(row=row_num, column=4, value=item.get('time_block', '')).alignment = center_align
        ws.cell(row=row_num, column=5, value=item.get('discipline', '')).alignment = left_align
        ws.cell(row=row_num, column=6, value=item.get('topic', '')).alignment = left_align
        ws.cell(row=row_num, column=7, value=item.get('instructor_name', '')).alignment = left_align
        ws.cell(row=row_num, column=8, value=item.get('location', '')).alignment = left_align
        ws.cell(row=row_num, column=9, value=item.get('hours', 0)).alignment = center_align

        # Применяем границы ко всем ячейкам строки
        for col in range(1, 10):
            ws.cell(row=row_num, column=col).border = thin_border

        row_num += 1

    # 🔹 Итого часов
    total_hours = sum(float(item.get('hours', 0)) for item in schedule)
    ws.merge_cells(f'A{row_num}:C{row_num}')
    ws.cell(row=row_num, column=1, value="ВСЕГО ЧАСОВ:").font = Font(bold=True, size=11)
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='right')
    ws.cell(row=row_num, column=9, value=total_hours).font = Font(bold=True, size=11)
    ws.cell(row=row_num, column=9).alignment = center_align

    # Границы для итоговой строки
    for col in range(1, 10):
        ws.cell(row=row_num, column=col).border = thin_border

    # 🔹 Авто-ширина колонок
    column_widths = [5, 12, 12, 18, 25, 35, 20, 20, 8]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 🔹 Сохраняем в байты
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.read()