# backend_doc_tya/api/utils/rauc_export.py
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.db.models import Prefetch
from api.models import Enrollment, Certificate

# Маппинг локаций из вашей модели
LOCATION_MAP = {
    "KJA": {"addr_id": "174", "dept_id": ""},
    "DME": {"addr_id": "176", "dept_id": "37"},
}


# backend_doc_tya/api/utils/rauc_export.py

def generate_rauc_excel(enrollments_qs):
    """Генерация Excel. Предполагает, что данные уже загружены с нужными prefetch."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лист1"

    headers = [
        'SURNAME', 'NAME', 'PATRONYMIC', 'DBIRTH', 'PROG_ID', 'MOD_ID',
        'DBEGINEXT', 'DBEGIN', 'DEND', 'NGROUP', 'ADDR_ID', 'DCAT_ID',
        'NDOC', 'DDOC', 'FPTITLE_ID', 'TPTITLE_ID', 'DENDDOC', 'DEPT_ID'
    ]
    ws.append(headers)

    # Стиль заголовков
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True, size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 🔹 НЕ делаем новых prefetch — используем уже загруженные данные
    for enr in enrollments_qs:  # ← данные уже с prefetch из вьюхи
        student = enr.student
        module = enr.module
        course = module.course
        loc = LOCATION_MAP.get(enr.location, {"addr_id": "", "dept_id": ""})

        # 🔹 Используем .all() или .first() на уже загруженных related managers
        cert = None
        for c in enr.certificates.all():  # ← не делает новый запрос!
            if c.cert_type == 'successful':
                cert = c
                break

        section_result = None
        for sr in enr.section_results.all():  # ← не делает новый запрос!
            if sr.completed_at:
                section_result = sr
                break
        # Или берём первый, если нет с датой
        if not section_result:
            section_result = enr.section_results.first()

        row = [
            student.surname if student.surname else '',
            student.name if student.name else '',
            student.patronymic if student.patronymic else '',
            student.date_of_birth.strftime('%d.%m.%Y') if student.date_of_birth else '',
            course.prog_id if course.prog_id else '',
            module.mod_id if module.mod_id else '',
            enr.enrolled_at.strftime('%d.%m.%Y') if enr.enrolled_at else '',
            enr.start_face_to_face.strftime('%d.%m.%Y') if enr.start_face_to_face else '',
            enr.completed_at.strftime('%d.%m.%Y') if enr.completed_at else '',
            section_result.ngroup if section_result and section_result.ngroup else f"{enr.group}-{enr.application}",
            loc['addr_id'],
            student.dcat_id if student.dcat_id else '',  # ✅ DCAT_ID теперь из модели
            cert.cert_number if cert and cert.cert_number else '',
            section_result.completed_at.strftime('%d.%m.%Y') if section_result and section_result.completed_at else '',
            cert.prepared_by.rauts_id if cert and cert.prepared_by and cert.prepared_by.rauts_id else '',
            cert.approved_by.rauts_id if cert and cert.approved_by and cert.approved_by.rauts_id else '',
            section_result.completed_at.strftime('%d.%m.%Y') if section_result and section_result.completed_at else '',
            loc['dept_id'],
        ]
        ws.append(row)

    # Авто-ширина и формат дат
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 25)

    date_columns = [4, 7, 8, 9, 14, 17]
    for row in ws.iter_rows(min_row=2):
        for idx, cell in enumerate(row, 1):
            if idx in date_columns and cell.value and isinstance(cell.value, str) and '.' in cell.value:
                cell.number_format = 'DD.MM.YYYY'

    ws.freeze_panes = 'A2'
    return wb